#!/usr/bin/env python
# -*- coding: utf-8 -*-
#pip install --user scikit-learn torch torchvision openpyxl xgboost lightgbm pysr physo eql
"""
Benchmark v7 with MONITOR logs (ASCII-safe).
- Nonlinear only: RandomForest + optional XGBoost, LightGBM
- Symbolic models disabled in v7
- Nested CV with RandomizedSearchCV
- Stacking meta-learner (ElasticNet on OOF predictions)
- Full tracebacks on build failures for easier debugging
- No imputation changes; no operator removal (Option 2 ONLY)
"""

import os, json, warnings, time, traceback, importlib
import argparse
import csv
from importlib.util import find_spec
from datetime import datetime
import numpy as np

from sklearn.model_selection import KFold, RandomizedSearchCV, cross_val_predict, BaseCrossValidator
from sklearn.compose import ColumnTransformer, TransformedTargetRegressor
from sklearn.preprocessing import OneHotEncoder, StandardScaler, QuantileTransformer

class SafeQuantileTransformer(QuantileTransformer):
    """QuantileTransformer that auto-caps n_quantiles to n_samples (prevents sklearn warnings in small-N CV)."""
    def fit(self, X, y=None):
        try:
            n = X.shape[0]
            if self.n_quantiles is None:
                self.n_quantiles = n
            else:
                self.n_quantiles = int(min(int(self.n_quantiles), int(n)))
        except Exception:
            pass
        return super().fit(X, y)
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.linear_model import ElasticNet, Ridge
from sklearn.neural_network import MLPRegressor
from sklearn.ensemble import RandomForestRegressor
from scipy.stats import uniform, randint, loguniform

warnings.filterwarnings("ignore", category=UserWarning)


# ======= ROUND-ROBIN CV =======
class RoundRobinKFold(BaseCrossValidator):
    """
    Deterministic round-robin K-fold splitter.

    Datapoint i (0-based) goes to fold (i % n_splits).
    Equivalent to 1-based rule: Fold = (N-1) % n_splits + 1.
    """
    def __init__(self, n_splits=5):
        if n_splits < 2:
            raise ValueError("n_splits must be at least 2.")
        self.n_splits = int(n_splits)

    def get_n_splits(self, X=None, y=None, groups=None):
        return self.n_splits

    def split(self, X, y=None, groups=None):
        n_samples = len(X)
        indices = np.arange(n_samples)
        fold_ids = indices % self.n_splits
        for fold in range(self.n_splits):
            test_idx = indices[fold_ids == fold]
            train_idx = indices[fold_ids != fold]
            yield train_idx, test_idx

START_TS = time.time()

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def elapsed():
    s = time.time() - START_TS
    h = int(s // 3600); m = int((s % 3600) // 60); sec = s % 60
    return f"{h:02d}:{m:02d}:{sec:05.2f}"

def log(msg):
    # ASCII-safe logging (avoid CP949 issues on Windows console)
    safe = (
        str(msg)
        .replace("—", "-").replace("–", "-").replace("±", "+/-")
        .replace("²", "^2").replace("³", "^3").replace("·", "*")
        .replace("×", "x").replace("√", "sqrt").replace("∑", "Sum")
        .replace("R²", "R2")
    )
    print(f"[MONITOR {now()} +{elapsed()}] {safe}", flush=True)

def sysinfo():
    info = {}
    try:
        import psutil
        vm = psutil.virtual_memory()
        info["ram_total_gb"] = round(vm.total / (1024**3), 2)
        info["ram_available_gb"] = round(vm.available / (1024**3), 2)
    except Exception:
        info["ram_total_gb"] = info["ram_available_gb"] = "n/a"
    info["cpu_count"] = os.cpu_count()
    return info

# ======= CONFIG =======
TARGET_METRIC = "Ea"
DATA_PATH  = TARGET_METRIC + ".xlsx"
OUT_DIR    = "."
TARGET_COL = TARGET_METRIC
ID_LIKE_COLS = ["CompositionName"]
N_SPLITS   = 5
RANDOM_STATE = 42
N_JOBS = max(1, (os.cpu_count() or 2) - 1)

# ======= DATA-SIZE REGIMES (hyperparameter presets) =======
# Regimes tuned for typical dataset sizes (n_samples):
# - n100:   ~100 datapoints
# - n500:   ~500 datapoints
# - n1000:  ~1000 datapoints
# - n5000:  ~5000 datapoints
REGIME = "auto"  # may be overridden via CLI --regime
N_SAMPLES = None  # populated after data load

def _select_regime(n_samples: int, override: str = "auto") -> str:
    if override and override.lower() != "auto":
        return override
    if n_samples <= 200:
        return "n100"
    if n_samples <= 800:
        return "n500"
    if n_samples <= 2000:
        return "n1000"
    return "n5000"

def _regime_summary(name: str) -> str:
    return {
        "n100":  "~100 datapoints (tiny)",
        "n500":  "~500 datapoints (small)",
        "n1000": "~1000 datapoints (medium)",
        "n5000": "~5000 datapoints (large)",
        "auto":  "auto (choose from n_samples)"
    }.get(name, name)
PYSR_PROFILE = os.environ.get("PYSR_PROFILE", "desktop")  # 'aggressive' for heavy search


# ======= CLI (model selection) =======
def parse_args(available_model_names):
    """Parse CLI args for selecting which models to run."""
    p = argparse.ArgumentParser(
        description="Benchmark runner with round-robin CV and selectable models."
    )
    g = p.add_mutually_exclusive_group()
    g.add_argument("--model", type=str, default=None,
                   help="Run a single model by name (case-insensitive).")
    g.add_argument("--models", type=str, default=None,
                   help="Comma-separated list of model names, or ALL.")
    p.add_argument("--list-models", action="store_true",
                   help="Print available model names and exit.")
    p.add_argument("--no-stacking", action="store_true",
                   help="Disable stacking ensemble (useful when running a single model).")
    p.add_argument("--n-jobs", type=int, default=None,
                   help="Override parallel job count (default: cpu_count-1).")
    p.add_argument("--out-dir", type=str, default=None,
                   help="Override output directory (default: current OUT_DIR).")
    p.add_argument("--regime", type=str, default="auto", choices=["auto","n100","n500","n1000","n5000"], help="Hyperparameter regime by dataset size. auto chooses based on n_samples.")
    p.add_argument("--outer-fold", type=int, default=0,
                   help="Run only this outer fold (1..N_SPLITS). 0 means run all folds.")
    p.add_argument("--merge-only", action="store_true",
                   help="Only merge saved fold predictions into full OOF and compute metrics.")
    p.add_argument("--merge-dir", type=str, default=None,
                   help="Directory containing fold NPZ files to merge (default: out-dir).")
    args = p.parse_args()

    if args.list_models:
        print("\n".join(available_model_names))
        raise SystemExit(0)

    # Normalize selection
    selected = None
    if args.model:
        selected = [args.model]
    elif args.models:
        if args.models.strip().upper() == "ALL":
            selected = list(available_model_names)
        else:
            selected = [s.strip() for s in args.models.split(",") if s.strip()]

    canon = {name.lower(): name for name in available_model_names}

    if selected is None:
        selected_names = list(available_model_names)
    else:
        selected_names = []
        unknown = []
        for s in selected:
            key = s.lower()
            if key in canon:
                selected_names.append(canon[key])
            else:
                unknown.append(s)
        if unknown:
            raise ValueError(
                "Unknown model name(s): %s\nAvailable: %s"
                % (", ".join(unknown), ", ".join(available_model_names))
            )

    # Apply overrides
    global N_JOBS, OUT_DIR, TARGET_METRIC, DATA_PATH, TARGET_COL, REGIME
    if args.n_jobs is not None:
        N_JOBS = int(args.n_jobs)
    if args.out_dir is not None:
        OUT_DIR = str(args.out_dir)
    target_metric = getattr(args, "target_metric", None)
    if target_metric is not None:
        TARGET_METRIC = str(target_metric)
        DATA_PATH = TARGET_METRIC + ".xlsx"
        TARGET_COL = TARGET_METRIC


    # Regime override (applied after data load when n_samples is known)
    REGIME = str(args.regime)

    return set(selected_names), args

# ======= OPTIONAL LIB FLAGS =======
HAS_XGB = find_spec("xgboost") is not None
HAS_LGBM = find_spec("lightgbm") is not None
HAS_PYSR = find_spec("pysr") is not None
HAS_PHYSO = find_spec("physo") is not None

# ======= EQL detection with ENV override + INTERNAL fallback =======
EQL_MODULE_ENV = os.environ.get("EQL_MODULE")  # e.g., "eqlx"
EQL_CLASS_ENV  = os.environ.get("EQL_CLASS")   # e.g., "EQLRegressor"

HAS_EQL = False
EQL_IMPORT_PATH = None
EQL_CLASS_NAME  = None
EQL_INTERNAL    = False  # if True, use the internal fallback class defined below

def resolve_eql():
    global HAS_EQL, EQL_IMPORT_PATH, EQL_CLASS_NAME, EQL_INTERNAL
    # 1) Honor explicit env override first
    if EQL_MODULE_ENV and EQL_CLASS_ENV:
        try:
            importlib.import_module(EQL_MODULE_ENV)
            HAS_EQL = True
            EQL_IMPORT_PATH = EQL_MODULE_ENV
            EQL_CLASS_NAME  = EQL_CLASS_ENV
            log(f"EQL resolved via env: module={EQL_IMPORT_PATH}, class={EQL_CLASS_NAME}")
            return
        except Exception as e:
            log(f"[WARN] Env-specified EQL import failed: {e}")

    # 2) Try common names
    for mod in ["eql", "eqlearn", "eql_pytorch", "eqltorch", "EQL"]:
        try:
            if find_spec(mod) is None:
                continue
            _m = importlib.import_module(mod)
            for klass in ["EQLRegressor", "EQL"]:
                if hasattr(_m, klass):
                    HAS_EQL = True
                    EQL_IMPORT_PATH = mod
                    EQL_CLASS_NAME  = klass
                    log(f"EQL resolved: module={EQL_IMPORT_PATH}, class={EQL_CLASS_NAME}")
                    return
        except Exception as e:
            log(f"[WARN] EQL probe for {mod} failed: {e}")

    # 3) INTERNAL fallback
    HAS_EQL = True
    EQL_INTERNAL = True
    log("EQL not found in env/common modules. Falling back to INTERNAL EQLRegressor.")

resolve_eql()

# ======= HELPERS =======
def rmse(y_true, y_pred):
    return float(np.sqrt(mean_squared_error(y_true, y_pred)))
def mae(y_true, y_pred):
    return float(mean_absolute_error(y_true, y_pred))
def metrics_dict(name, y_true, y_pred):
    return {"Model": name, "RMSE": rmse(y_true, y_pred), "MAE": mae(y_true, y_pred), "R2": float(r2_score(y_true, y_pred))}

def _fold_npz_path(model_name, fold, out_dir=OUT_DIR, metric=TARGET_METRIC):
    os.makedirs(os.path.join(out_dir, "fold_artifacts"), exist_ok=True)
    return os.path.join(out_dir, "fold_artifacts", f"{metric}__{model_name}__fold{fold}.npz")

def save_fold_artifact(model_name, fold, te_idx, y_true, y_pred, out_dir=OUT_DIR):
    fn = _fold_npz_path(model_name, fold, out_dir=out_dir)
    np.savez_compressed(fn,
                        model=str(model_name),
                        fold=int(fold),
                        te_idx=np.asarray(te_idx, dtype=int),
                        y_true=np.asarray(y_true, dtype=float),
                        y_pred=np.asarray(y_pred, dtype=float))
    log(f"Saved fold artifact: {fn}")

def merge_fold_artifacts(model_name, n_samples, out_dir=OUT_DIR, metric=TARGET_METRIC):
    """Merge fold artifacts into a full OOF vector."""
    oof = np.full(n_samples, np.nan, dtype=float)
    for fold in range(1, N_SPLITS + 1):
        fn = _fold_npz_path(model_name, fold, out_dir=out_dir, metric=metric)
        if not os.path.exists(fn):
            raise FileNotFoundError(f"Missing fold artifact: {fn}")
        z = np.load(fn, allow_pickle=True)
        te_idx = z["te_idx"].astype(int)
        y_pred = z["y_pred"].astype(float)
        oof[te_idx] = y_pred
    if np.isnan(oof).any():
        raise RuntimeError(f"OOF merge incomplete for {model_name}: still has NaNs")
    return oof


def save_parity_fold(model_name, fold, y_true_fold, y_pred_fold, out_dir=OUT_DIR, metric=TARGET_METRIC):
    """
    Save parity CSV for a single outer fold (y_true/y_pred only for that fold).
    This is safe for fold-parallel runs where full-length OOF is incomplete.
    """
    os.makedirs(out_dir, exist_ok=True)
    fn = os.path.join(out_dir, f"{metric}_predictions_{model_name}_fold{fold}.csv")
    with open(fn, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([f"True_{metric}", f"Pred_{metric}"])
        for yt, yp in zip(y_true_fold, y_pred_fold):
            w.writerow([float(yt), float(yp)])
    log(f"Saved fold parity CSV: {fn}")

def save_parity(name, y_true, y_pred, out_dir=OUT_DIR):
    """Save true vs predicted values to CSV without pandas."""
    fn = os.path.join(out_dir, f"{TARGET_METRIC}_predictions_{name}.csv")
    y_true = np.asarray(y_true).ravel()
    y_pred = np.asarray(y_pred).ravel()
    arr = np.column_stack([y_true, y_pred])
    os.makedirs(out_dir, exist_ok=True)
    np.savetxt(fn, arr, delimiter=",", header="True_Ea,Pred_Ea", comments="", fmt="%.10g")
    log(f"Saved parity CSV: {fn}")

def build_preprocessor(X):
    """Build a preprocessor for either a pandas DataFrame or a numpy array.

    This script is intended to run on HPC environments where pandas may be unusable
    (e.g., Python built without _bz2). If X is a numpy array, all columns are treated
    as numeric and standardized.
    """
    try:
        import pandas as pd  # optional
    except Exception:
        pd = None

    if pd is not None and hasattr(pd, "DataFrame") and isinstance(X, pd.DataFrame):
        cat_cols = X.select_dtypes(include=["object", "category"]).columns.tolist()
        num_cols = X.select_dtypes(exclude=["object", "category"]).columns.tolist()
        try:
            ohe = OneHotEncoder(handle_unknown="ignore", sparse_output=False)
        except TypeError:
            ohe = OneHotEncoder(handle_unknown="ignore", sparse=False)
        preprocessor = ColumnTransformer([
            ("cat", ohe, cat_cols),
            ("num", StandardScaler(), num_cols),
        ], remainder="drop", verbose_feature_names_out=False)
        log(f"Preprocessor configured | num={len(num_cols)} cat={len(cat_cols)}")
        return preprocessor

    n_features = int(X.shape[1])
    preprocessor = ColumnTransformer(
        [("num", StandardScaler(), list(range(n_features)))],
        remainder="drop",
        verbose_feature_names_out=False,
    )
    log(f"Preprocessor configured | n_features={n_features} num={n_features} cat=0")
    return preprocessor
def wrap_ttr(estimator):
    return TransformedTargetRegressor(
        regressor=estimator,
        transformer=SafeQuantileTransformer(n_quantiles=1000, output_distribution="normal", random_state=RANDOM_STATE)
    )

# ======= PARAM SPACES (NOTE: 'regressor__regressor__*' keys!) =======
def space_elastic():
    r = _select_regime(N_SAMPLES or 1000, REGIME)
    # For small-N, avoid very small alpha (overfitting / instability).
    if r == "n100":
        return {
            "regressor__regressor__alpha": loguniform(1e-3, 1e+1),
            "regressor__regressor__l1_ratio": uniform(0.05, 0.95),
        }
    if r == "n500":
        return {
            "regressor__regressor__alpha": loguniform(1e-4, 1e+1),
            "regressor__regressor__l1_ratio": uniform(0.01, 0.99),
        }
    if r == "n1000":
        return {
            "regressor__regressor__alpha": loguniform(1e-5, 1e+1),
            "regressor__regressor__l1_ratio": uniform(0.0, 1.0),
        }
    # n5000
    return {
        "regressor__regressor__alpha": loguniform(1e-6, 1e+1),
        "regressor__regressor__l1_ratio": uniform(0.0, 1.0),
    }
def base_elastic():
    return ElasticNet(max_iter=20000, random_state=RANDOM_STATE)

def space_ridge():
    r = _select_regime(N_SAMPLES or 1000, REGIME)
    # Ridge is stable; regimes mainly adjust typical alpha scale.
    if r == "n100":
        return {"regressor__regressor__alpha": loguniform(1e-2, 1e4)}
    if r == "n500":
        return {"regressor__regressor__alpha": loguniform(1e-3, 1e4)}
    if r == "n1000":
        return {"regressor__regressor__alpha": loguniform(1e-4, 1e4)}
    # n5000
    return {"regressor__regressor__alpha": loguniform(1e-5, 1e4)}
def base_ridge():
    return Ridge(max_iter=20000)

def space_rf():

    r = (_select_regime(N_SAMPLES or 1000, REGIME))
    if r == "n100":
        return {
            "regressor__regressor__n_estimators": randint(200, 900),
            "regressor__regressor__max_depth": randint(2, 14),
            "regressor__regressor__min_samples_split": randint(2, 10),
            "regressor__regressor__min_samples_leaf": randint(1, 12),
            "regressor__regressor__max_features": uniform(0.4, 0.6),
        }
    if r == "n500":
        return {
            "regressor__regressor__n_estimators": randint(400, 1400),
            "regressor__regressor__max_depth": randint(3, 22),
            "regressor__regressor__min_samples_split": randint(2, 12),
            "regressor__regressor__min_samples_leaf": randint(1, 10),
            "regressor__regressor__max_features": uniform(0.3, 0.7),
        }
    if r == "n1000":
        return {
            "regressor__regressor__n_estimators": randint(600, 1800),
            "regressor__regressor__max_depth": randint(4, 30),
            "regressor__regressor__min_samples_split": randint(2, 14),
            "regressor__regressor__min_samples_leaf": randint(1, 8),
            "regressor__regressor__max_features": uniform(0.3, 0.7),
        }
    # n5000
    return {
        "regressor__regressor__n_estimators": randint(900, 2600),
        "regressor__regressor__max_depth": randint(4, 40),
        "regressor__regressor__min_samples_split": randint(2, 16),
        "regressor__regressor__min_samples_leaf": randint(1, 6),
        "regressor__regressor__max_features": uniform(0.2, 0.8),
    }
def base_rf():
    return RandomForestRegressor(random_state=RANDOM_STATE, n_jobs=1, bootstrap=True)

def space_mlp():

    r = (_select_regime(N_SAMPLES or 1000, REGIME))
    if r == "n100":
        return {
            "regressor__regressor__hidden_layer_sizes": [(64,32), (128,64), (64,64,32)],
            "regressor__regressor__activation": ["relu", "tanh"],
            "regressor__regressor__alpha": loguniform(1e-5, 5e-1),
            "regressor__regressor__learning_rate_init": loguniform(1e-4, 5e-3),
            "regressor__regressor__early_stopping": [True],
            "regressor__regressor__max_iter": [2000],
            "regressor__regressor__verbose": [False],
        }
    if r == "n500":
        return {
            "regressor__regressor__hidden_layer_sizes": [(128,64), (256,128), (256,128,64)],
            "regressor__regressor__activation": ["relu", "tanh"],
            "regressor__regressor__alpha": loguniform(1e-6, 1e-2),
            "regressor__regressor__learning_rate_init": loguniform(3e-4, 3e-2),
            "regressor__regressor__early_stopping": [True],
            "regressor__regressor__max_iter": [3000],
            "regressor__regressor__verbose": [False],
        }
    if r == "n1000":
        return {
            "regressor__regressor__hidden_layer_sizes": [(256,128), (256,128,64), (512,256,128)],
            "regressor__regressor__activation": ["relu", "tanh"],
            "regressor__regressor__alpha": loguniform(1e-6, 1e-2),
            "regressor__regressor__learning_rate_init": loguniform(3e-4, 3e-2),
            "regressor__regressor__early_stopping": [True],
            "regressor__regressor__max_iter": [4000],
            "regressor__regressor__verbose": [False],
        }
    # n5000
    return {
        "regressor__regressor__hidden_layer_sizes": [(256,128), (512,256,128), (512,256,256,128)],
        "regressor__regressor__activation": ["relu", "tanh"],
        "regressor__regressor__alpha": loguniform(1e-7, 1e-3),
        "regressor__regressor__learning_rate_init": loguniform(1e-4, 1e-2),
        "regressor__regressor__early_stopping": [True],
        "regressor__regressor__max_iter": [6000],
        "regressor__regressor__verbose": [False],
    }
def base_mlp():
    return MLPRegressor(random_state=RANDOM_STATE, solver="adam", verbose=False)

def space_xgb():

    r = (_select_regime(N_SAMPLES or 1000, REGIME))
    if r == "n100":
        return {
            "regressor__regressor__n_estimators": randint(200, 1500),
            "regressor__regressor__learning_rate": loguniform(0.02, 0.3),
            "regressor__regressor__max_depth": randint(2, 8),
            "regressor__regressor__subsample": uniform(0.7, 0.3),
            "regressor__regressor__colsample_bytree": uniform(0.6, 0.4),
            "regressor__regressor__min_child_weight": randint(1, 8),
            "regressor__regressor__reg_lambda": loguniform(0.5, 50.0),
            "regressor__regressor__gamma": loguniform(1e-9, 1e-1),
        }
    if r == "n500":
        return {
            "regressor__regressor__n_estimators": randint(600, 2500),
            "regressor__regressor__learning_rate": loguniform(0.01, 0.25),
            "regressor__regressor__max_depth": randint(3, 10),
            "regressor__regressor__subsample": uniform(0.6, 0.4),
            "regressor__regressor__colsample_bytree": uniform(0.5, 0.5),
            "regressor__regressor__min_child_weight": randint(1, 10),
            "regressor__regressor__reg_lambda": loguniform(0.5, 80.0),
            "regressor__regressor__gamma": loguniform(1e-9, 1e-1),
        }
    if r == "n1000":
        return {
            "regressor__regressor__n_estimators": randint(900, 3200),
            "regressor__regressor__learning_rate": loguniform(0.01, 0.2),
            "regressor__regressor__max_depth": randint(3, 12),
            "regressor__regressor__subsample": uniform(0.6, 0.4),
            "regressor__regressor__colsample_bytree": uniform(0.5, 0.5),
            "regressor__regressor__min_child_weight": randint(1, 12),
            "regressor__regressor__reg_lambda": loguniform(1.0, 120.0),
            "regressor__regressor__gamma": loguniform(1e-9, 1e-1),
        }
    # n5000
    return {
        "regressor__regressor__n_estimators": randint(1200, 6000),
        "regressor__regressor__learning_rate": loguniform(0.01, 0.15),
        "regressor__regressor__max_depth": randint(3, 14),
        "regressor__regressor__subsample": uniform(0.6, 0.4),
        "regressor__regressor__colsample_bytree": uniform(0.4, 0.6),
        "regressor__regressor__min_child_weight": randint(1, 20),
        "regressor__regressor__reg_lambda": loguniform(1.0, 200.0),
        "regressor__regressor__gamma": loguniform(1e-9, 2e-1),
    }
def base_xgb():
    from xgboost import XGBRegressor  # local import to avoid NameError
    return XGBRegressor(
        objective="reg:squarederror",
        tree_method="hist",
        random_state=RANDOM_STATE,
        n_jobs=1,
        verbosity=1,
        max_cat_to_onehot=64
    )

def space_lgbm():

    r = (_select_regime(N_SAMPLES or 1000, REGIME))
    if r == "n100":
        return {
            "regressor__regressor__n_estimators": randint(300, 1600),
            "regressor__regressor__learning_rate": loguniform(0.02, 0.3),
            "regressor__regressor__num_leaves": randint(7, 64),
            "regressor__regressor__min_child_samples": randint(1, 25),
            "regressor__regressor__feature_fraction": uniform(0.7, 0.3),
            "regressor__regressor__bagging_fraction": uniform(0.7, 0.3),
            "regressor__regressor__lambda_l1": loguniform(1e-6, 5.0),
            "regressor__regressor__lambda_l2": loguniform(1e-6, 5.0),
        }
    if r == "n500":
        return {
            "regressor__regressor__n_estimators": randint(600, 2600),
            "regressor__regressor__learning_rate": loguniform(0.01, 0.25),
            "regressor__regressor__num_leaves": randint(15, 128),
            "regressor__regressor__min_child_samples": randint(3, 90),
            "regressor__regressor__feature_fraction": uniform(0.6, 0.4),
            "regressor__regressor__bagging_fraction": uniform(0.6, 0.4),
            "regressor__regressor__lambda_l1": loguniform(1e-6, 10.0),
            "regressor__regressor__lambda_l2": loguniform(1e-6, 10.0),
        }
    if r == "n1000":
        return {
            "regressor__regressor__n_estimators": randint(900, 3800),
            "regressor__regressor__learning_rate": loguniform(0.01, 0.2),
            "regressor__regressor__num_leaves": randint(31, 256),
            "regressor__regressor__min_child_samples": randint(5, 140),
            "regressor__regressor__feature_fraction": uniform(0.6, 0.4),
            "regressor__regressor__bagging_fraction": uniform(0.6, 0.4),
            "regressor__regressor__lambda_l1": loguniform(1e-6, 20.0),
            "regressor__regressor__lambda_l2": loguniform(1e-6, 20.0),
        }
    # n5000
    return {
        "regressor__regressor__n_estimators": randint(1200, 6500),
        "regressor__regressor__learning_rate": loguniform(0.01, 0.15),
        "regressor__regressor__num_leaves": randint(31, 512),
        "regressor__regressor__min_child_samples": randint(10, 220),
        "regressor__regressor__feature_fraction": uniform(0.5, 0.5),
        "regressor__regressor__bagging_fraction": uniform(0.5, 0.5),
        "regressor__regressor__lambda_l1": loguniform(1e-6, 50.0),
        "regressor__regressor__lambda_l2": loguniform(1e-6, 50.0),
    }
def base_lgbm():
    from lightgbm import LGBMRegressor  # local import
    return LGBMRegressor(random_state=RANDOM_STATE, n_jobs=1, verbose=-1)

# ======= SYMBOLIC MODELS =======
def build_pysr():
    from pysr import PySRRegressor
    if PYSR_PROFILE.lower().startswith("aggr"):
        niterations = 500; populations = 32; maxsize = 50
    else:
        niterations = 220; populations = 12; maxsize = 40
    kwargs = dict(
        niterations=niterations,
        binary_operators=["+", "-", "*", "/", "pow"],
        unary_operators=["exp", "log", "sin", "cos", "tan", "sqrt", "abs"],
        maxsize=maxsize,
        procs=N_JOBS,
        model_selection="best",
        elementwise_loss="L2DistLoss()",
        populations=populations,
        nested_constraints={"pow": {"pow": 0}},
        random_state=RANDOM_STATE,
        progress=True,
    )
    try:
        return PySRRegressor(**kwargs)
    except TypeError:
        kwargs.pop("elementwise_loss", None)
        kwargs["loss"] = "L2DistLoss()"
        return PySRRegressor(**kwargs)

# --- PhySO sklearn-style wrapper around physo.SR ---
class PhysoSklearnWrapper:
    """Sklearn-style adapter around physo.SR(...). Uses SAFE math for eval."""
    def __init__(self, epochs=10, op_names=None, parallel_mode=False, n_cpus=None, verbose=True):
        self.epochs = epochs
        self.op_names = op_names or ["mul","add","sub","div","n2","sqrt","neg","exp","log","sin","cos"]
        self.parallel_mode = parallel_mode
        self.n_cpus = n_cpus
        self.verbose = verbose
        self._expr = None
        self._xnames = None
        self._func = None

    def fit(self, X, y):
        import physo
        import sympy as sp
        import physo.learn.monitoring as monitoring
        import numpy as _np

        X = np.asarray(X)
        y = np.asarray(y).astype(float)
        if X.ndim != 2:
            raise ValueError("PhySO expects X as 2D array (n_samples, n_features).")
        X_phy = X.T  # (n_features, n_samples)
        self._xnames = [f"x{i}" for i in range(X.shape[1])]

        run_logger     = lambda : monitoring.RunLogger(save_path=None, do_save=False)
        run_visualiser = lambda : monitoring.RunVisualiser(epoch_refresh_rate=10, save_path=None, do_show=False, do_prints=self.verbose, do_save=False)

        # Config fallback
        try:
            run_cfg = getattr(importlib.import_module("physo.config.config0"), "config0")
        except Exception:
            run_cfg = None

        log("PhySO SR() starting...")
        expr, logs = physo.SR(
            X_phy, y,
            X_names=self._xnames,
            y_name="target",
            op_names=self.op_names,
            get_run_logger=run_logger,
            get_run_visualiser=run_visualiser,
            run_config=run_cfg,
            parallel_mode=self.parallel_mode,
            n_cpus=self.n_cpus,
            epochs=self.epochs
        )
        self._expr = expr

        # Build a safe numpy callable via sympy
        sym = expr.get_infix_sympy()
        symbols = [sp.symbols(n) for n in self._xnames]

        # ---- SAFE math shims (Option 2) ----
        def _safe_log(x):
            x = _np.asarray(x)
            with _np.errstate(all="ignore"):
                return _np.log(_np.abs(x) + 1e-12)
        def _safe_pow(a, b):
            a = _np.asarray(a); b = _np.asarray(b)
            with _np.errstate(all="ignore"):
                return _np.power(_np.abs(a), b)

        self._func = sp.lambdify(
            symbols,
            sym,
            modules=[{"log": _safe_log, "pow": _safe_pow}, "numpy"]
        )

        try:
            pretty = expr.get_infix_pretty()
        except Exception:
            pretty = str(sym)
        log("PhySO expression: " + pretty)
        return self

    def predict(self, X):
        if self._func is None:
            raise RuntimeError("PhySO wrapper not fitted.")
        import numpy as _np
        X = _np.asarray(X)
        cols = [X[:, i] for i in range(X.shape[1])]
        with _np.errstate(all="ignore"):
            yhat = self._func(*cols)
        yhat = _np.asarray(yhat, dtype=float)

        # Scrub non-finite outputs
        finite = _np.isfinite(yhat)
        if not finite.all():
            n_bad = int((~finite).sum())
            repl = _np.nanmedian(yhat[finite]) if finite.any() else 0.0
            yhat = yhat.copy()
            yhat[~finite] = repl
            log(f"[WARN] PhySO predict produced {n_bad} non-finite values; replaced with median={repl:.6g}")
        return yhat

def build_physo():
    if not HAS_PHYSO:
        raise ImportError("physo not found")
    r = _select_regime(N_SAMPLES or 1000, REGIME)
    # Scale training budget by dataset size. Keep conservative for small-N to avoid overfit and long runs.
    epochs_map = {"n100": 10, "n500": 10, "n1000": 10, "n5000": 10}
    epochs = int(epochs_map.get(r, 200))
    # Optional override via env (kept for backward compatibility)
    if PYSR_PROFILE.lower().startswith("aggr"):
        epochs = max(epochs, int(1.5 * epochs))
    return PhysoSklearnWrapper(epochs=epochs, parallel_mode=False, verbose=True)

# --- INTERNAL EQL fallback (sklearn-style) ---
import numpy as np
import torch
import torch.nn as nn

class _InternalEQLRegressor:
    """
    Internal EQL-style regressor with Top-K masking on the output layer.
    """
    def __init__(
        self,
        max_terms=20,
        hidden_units=None,
        mult_units=None,
        epochs=600,
        lr=1e-3,
        l1=1e-3,
        seed=42,
        verbose=True,
        random_state=None,  # accepted for compatibility
    ):
        self.max_terms = int(max_terms)
        self.hidden_units = None if hidden_units is None else int(hidden_units)
        self.mult_units = None if mult_units is None else int(mult_units)
        self.epochs = int(epochs)
        self.lr = float(lr)
        self.l1 = float(l1)
        self.seed = int(seed if random_state is None else random_state)
        self.verbose = bool(verbose)
        self._model = None
        self._n_in = None

    def _choose_capacity(self, n_in):
        if self.hidden_units is None or self.mult_units is None:
            if self.max_terms < 9:
                hidden = 1
                mult   = max(0, self.max_terms - 8)
            else:
                approx_hidden = max(1, min((self.max_terms // 2) // 8, 6))
                used = 8 * approx_hidden
                remaining = max(0, self.max_terms - used)
                hidden = approx_hidden
                mult   = remaining
        else:
            hidden = self.hidden_units
            mult   = self.mult_units
        total = 8*hidden + mult
        if total > self.max_terms:
            over = total - self.max_terms
            dec = min(mult, over)
            mult -= dec
            over -= dec
            while over > 0 and hidden > 1:
                hidden -= 1
                over -= 8
        assert 8*hidden + mult <= self.max_terms, "capacity over budget"
        return hidden, mult

    class _EQLNet(nn.Module):
        def __init__(self, n_in, hidden_units, mult_units):
            super().__init__()
            self.W_lin = nn.Linear(n_in, hidden_units, bias=True)
            self.W_sin = nn.Linear(n_in, hidden_units, bias=True)
            self.W_cos = nn.Linear(n_in, hidden_units, bias=True)
            self.W_sq  = nn.Linear(n_in, hidden_units, bias=True)
            self.W_cu  = nn.Linear(n_in, hidden_units, bias=True)
            self.W_exp = nn.Linear(n_in, hidden_units, bias=True)
            self.W_log = nn.Linear(n_in, hidden_units, bias=True)
            self.W_erf = nn.Linear(n_in, hidden_units, bias=True)
            self.P = nn.Linear(n_in, mult_units, bias=True)
            self.Q = nn.Linear(n_in, mult_units, bias=True)
            total_features = 8*hidden_units + mult_units
            self.out = nn.Linear(total_features, 1, bias=True)

        def forward(self, x):
            z_lin = self.W_lin(x)
            z_sin = torch.sin(self.W_sin(x))
            z_cos = torch.cos(self.W_cos(x))
            z_sq  = torch.square(self.W_sq(x))
            z_cu  = torch.pow(self.W_cu(x), 3)
            z_exp = torch.exp(self.W_exp(x))
            z_log = torch.log(torch.abs(self.W_log(x)) + 1e-6)
            z_erf = torch.erf(self.W_erf(x))
            z_mul = self.P(x) * self.Q(x)
            z = torch.cat([z_lin, z_sin, z_cos, z_sq, z_cu, z_exp, z_log, z_erf, z_mul], dim=1)
            return self.out(z)

    def _apply_topk(self, m):
        with torch.no_grad():
            w = m.out.weight.data.flatten()
            k = min(self.max_terms, w.numel())
            if k <= 0:
                return
            top_idx = torch.topk(w.abs(), k).indices
            mask = torch.zeros_like(w)
            mask[top_idx] = 1.0
            m.out.weight.data = (w * mask).reshape_as(m.out.weight)
            if m.out.weight.grad is not None:
                g = m.out.weight.grad.data.flatten()
                m.out.weight.grad.data = (g * mask).reshape_as(m.out.weight)

    def fit(self, X, y):
        torch.manual_seed(self.seed)
        np.random.seed(self.seed)
        X = np.asarray(X, dtype=np.float32)
        y = np.asarray(y, dtype=np.float32).reshape(-1, 1)
        assert X.ndim == 2 and X.shape[0] == y.shape[0]
        n_in = X.shape[1]
        self._n_in = n_in
        hidden, mult = self._choose_capacity(n_in)
        self.hidden_units_ = hidden
        self.mult_units_ = mult
        model = self._EQLNet(n_in, hidden, mult)
        opt = torch.optim.Adam(model.parameters(), lr=self.lr)
        loss_fn = nn.MSELoss()
        Xt = torch.from_numpy(X)
        yt = torch.from_numpy(y)
        for ep in range(self.epochs):
            model.train()
            opt.zero_grad()
            yhat = model(Xt)
            mse = loss_fn(yhat, yt)
            l1_pen = torch.norm(model.out.weight, p=1)
            l1_proj = (
                torch.norm(model.W_lin.weight,1) +
                torch.norm(model.W_sin.weight,1) +
                torch.norm(model.W_cos.weight,1) +
                torch.norm(model.W_sq.weight,1)  +
                torch.norm(model.P.weight,1) +
                torch.norm(model.Q.weight,1)
            )
            loss = mse + self.l1 * (l1_pen + 0.1*l1_proj)
            loss.backward()
            opt.step()
            self._apply_topk(model)
            if self.verbose and (ep % 100 == 0 or ep == self.epochs - 1):
                with torch.no_grad():
                    mse_val = loss_fn(model(Xt), yt).item()
                print(f"[Internal EQLRegressor] epoch {ep} MSE={mse_val:.6f} | K<={self.max_terms}")
        self._model = model.eval()
        return self

    def predict(self, X):
        if self._model is None:
            raise RuntimeError("Call fit before predict.")
        X = np.asarray(X, dtype=np.float32)
        Xt = torch.from_numpy(X)
        with torch.no_grad():
            yhat = self._model(Xt).cpu().numpy().flatten()
        return yhat

def build_eql():
    if EQL_INTERNAL:
        r = _select_regime(N_SAMPLES or 1000, REGIME)
        # Scale capacity + training budget with dataset size.
        cfg = {
            "n100":  {"max_terms": 12, "epochs": 600,  "lr": 2e-3, "l1": 3e-3},
            "n500":  {"max_terms": 20, "epochs": 1200, "lr": 1e-3, "l1": 2e-3},
            "n1000": {"max_terms": 30, "epochs": 2000, "lr": 8e-4, "l1": 1e-3},
            "n5000": {"max_terms": 60, "epochs": 5000, "lr": 5e-4, "l1": 5e-4},
        }.get(r, {"max_terms": 30, "epochs": 2000, "lr": 8e-4, "l1": 1e-3})
        log(f"Using INTERNAL EQLRegressor fallback | regime={r} | cfg={cfg}")
        return _InternalEQLRegressor(random_state=RANDOM_STATE, **cfg)
    # external path
    mod = importlib.import_module(EQL_IMPORT_PATH)
    cls = getattr(mod, EQL_CLASS_NAME)
    try:
        return cls(random_state=RANDOM_STATE)
    except TypeError:
        return cls()

def get_symbolic_specs(preprocessor):
    specs = []
    # PySR removed (Julia/PythonCall dependency not supported on this cluster)
    if HAS_PYSR:
        try:
            pysr = Pipeline([("prep", preprocessor), ("regressor", build_pysr())])
            specs.append(("PySR", pysr, None, 0))
            log("PySR enabled")
        except Exception as e:
            log(f"[WARN] PySR present but failed to build: {e}")
            traceback.print_exc()
    else:
        log("PySR not installed; skipping.")


    # PhySO (wrapper, SAFE eval)
    if HAS_PHYSO:
        try:
            physo = Pipeline([("prep", preprocessor), ("regressor", build_physo())])
            specs.append(("PhySO", physo, None, 0))
            log("PhySO enabled via wrapper (SAFE eval)")
        except Exception as e:
            log(f"[WARN] PhySO present but failed to build (wrapper): {e}")
            traceback.print_exc()
    else:
        log("PhySO not installed; skipping.")

    # EQL (env-aware + internal fallback)
    if HAS_EQL:
        try:
            eql = Pipeline([("prep", preprocessor), ("regressor", build_eql())])
            specs.append(("EQL", eql, None, 0))
            src = f"{'INTERNAL' if EQL_INTERNAL else EQL_IMPORT_PATH + '.' + EQL_CLASS_NAME}"
            log(f"EQL enabled: {src}")
        except Exception as e:
            log(f"[WARN] EQL detected but failed to build: {e}")
            log("Hint: Set EQL_MODULE and EQL_CLASS env vars to the correct module/class for your installation, or rely on the INTERNAL fallback.")
            traceback.print_exc()
    else:
        log("EQL not installed or no estimator found; INTERNAL fallback should have taken over.")
    return specs

# ======= CLASSIC MODELS =======

def make_model_space_list(preprocessor):
    models = []
    r = _select_regime(N_SAMPLES or 1000, REGIME)
    # Rough guidance for RandomizedSearchCV budget (per outer fold)
    n_iter_map = {
        "n100":  {"linear": 25, "rf": 20,  "mlp": 30, "xgb": 25,  "lgbm": 25},
        "n500":  {"linear": 40, "rf": 30,  "mlp": 50, "xgb": 40,  "lgbm": 40},
        "n1000": {"linear": 60, "rf": 40,  "mlp": 70, "xgb": 60,  "lgbm": 60},
        "n5000": {"linear": 80, "rf": 60,  "mlp": 90, "xgb": 90,  "lgbm": 90},
    }[r]
    log(f"Hyperparameter regime: {r} ({_regime_summary(r)}) | n_samples={N_SAMPLES}")

    # Linear models
    enet = Pipeline([("prep", preprocessor), ("regressor", base_elastic())])
    models.append(("ElasticNet", wrap_ttr(enet), space_elastic(), n_iter_map["linear"]))

    ridge = Pipeline([("prep", preprocessor), ("regressor", base_ridge())])
    models.append(("Ridge", wrap_ttr(ridge), space_ridge(), n_iter_map["linear"]))

    # Nonlinear models
    rf = Pipeline([("prep", preprocessor), ("regressor", base_rf())])
    models.append(("RandomForest", wrap_ttr(rf), space_rf(), n_iter_map["rf"]))

    mlp = Pipeline([("prep", preprocessor), ("regressor", base_mlp())])
    models.append(("MLP", wrap_ttr(mlp), space_mlp(), n_iter_map["mlp"]))

    if HAS_XGB:
        xgb = Pipeline([("prep", preprocessor), ("regressor", base_xgb())])
        models.append(("XGBoost", wrap_ttr(xgb), space_xgb(), n_iter_map["xgb"]))

    if HAS_LGBM:
        lgbm = Pipeline([("prep", preprocessor), ("regressor", base_lgbm())])
        models.append(("LightGBM", wrap_ttr(lgbm), space_lgbm(), n_iter_map["lgbm"]))

    return models
# ======= NESTED CV =======
def nested_cv_oof(X, y, model, param_space, model_name, n_splits=5, n_iter=50, random_state=RANDOM_STATE, outer_fold=0, save_fold_npz=True):
    log(f"Begin nested CV for {model_name} | outer folds={n_splits} | inner iters={n_iter if param_space is not None else 'n/a'}")
    outer = RoundRobinKFold(n_splits=n_splits)
    oof_pred = np.zeros(len(y), dtype=float)
    per_fold_params = []
    for fold, (tr_idx, te_idx) in enumerate(outer.split(X), 1):
        if outer_fold and int(outer_fold) != int(fold):
            continue
        log(f"[{model_name}] FOLD {fold}/{n_splits} | train={len(tr_idx)} test={len(te_idx)}")
        X_tr, X_te = X[tr_idx], X[te_idx]
        y_tr = y[tr_idx]
        t0 = time.time()
        try:
            if param_space is None:
                log(f"[{model_name}] Direct fit (internal search) starting...")
                est = model
                est.fit(X_tr, y_tr)
                y_hat = est.predict(X_te)
                per_fold_params.append({"internal_search": True})
            else:
                log(f"[{model_name}] RandomizedSearchCV starting...")
                search = RandomizedSearchCV(
                    estimator=model,
                    param_distributions=param_space,
                    n_iter=n_iter,
                    scoring="r2",
                    cv=RoundRobinKFold(n_splits=4),
                    n_jobs=N_JOBS,
                    random_state=random_state + fold,
                    verbose=2,
                    refit=True
                )
                search.fit(X_tr, y_tr)
                best_params = dict(search.best_params_)
                best_score = float(search.best_score_)
                log(f"[{model_name}] INNER-CV best_score={best_score:.6f}")
                log(f"[{model_name}] INNER-CV best_params={best_params}")
                # Save per-fold best params for reuse (Option C)
                try:
                    os.makedirs(os.path.join(OUT_DIR, 'best_params'), exist_ok=True)
                    with open(os.path.join(OUT_DIR, 'best_params', f'{model_name}_fold{fold}.json'), 'w') as f:
                        json.dump({'model': model_name, 'fold': fold, 'best_score': best_score, 'best_params': best_params}, f, indent=2)
                except Exception as _e:
                    log(f"[WARN] Failed to save best params JSON: {_e}")
                best_est = search.best_estimator_
                per_fold_params.append(search.best_params_)
                y_hat = best_est.predict(X_te)
            dt = time.time() - t0
            fold_r2 = r2_score(y[te_idx], y_hat)
            fold_rmse = rmse(y[te_idx], y_hat)
            fold_mae = mae(y[te_idx], y_hat)
            log(f"[{model_name}] FOLD {fold} done in {dt:.1f}s | R2={fold_r2:.4f} | RMSE={fold_rmse:.4f} | MAE={fold_mae:.4f}")
        except Exception as e:
            log(f"[ERROR] {model_name} FOLD {fold} failed: {e}")
            traceback.print_exc()
            raise
        if save_fold_npz:
            save_fold_artifact(model_name, fold, te_idx, y[te_idx], y_hat, out_dir=OUT_DIR)
        # If running a single outer fold, also emit per-fold parity CSV (40 CSVs is OK)
        if outer_fold and int(outer_fold) == int(fold):
            save_parity_fold(model_name, fold, y[te_idx], y_hat, out_dir=OUT_DIR, metric=TARGET_METRIC)
        oof_pred[te_idx] = y_hat
    return oof_pred, per_fold_params

def fit_full_with_best_params(model, best_params_list, name=""):
    from collections import Counter
    flat = {}
    for d in best_params_list:
        for k,v in d.items():
            flat.setdefault(k, []).append(v)
    agg = {}
    for k, vals in flat.items():
        v0 = vals[0]
        if isinstance(v0, (int, float, np.floating, np.integer)):
            try:
                med = float(np.median(vals))
                if any(s in k for s in ["n_estimators","max_depth","min_child","num_leaves"]):
                    med = int(round(med))
                agg[k] = med
            except Exception:
                agg[k] = v0
        else:
            agg[k] = Counter(vals).most_common(1)[0][0]
    try:
        model.set_params(**agg)
        log(f"[{name}] Aggregated best params applied: {agg}")
    except Exception as e:
        log(f"[{name}] Could not apply aggregated params: {e}")
    return model, agg

def main(selected_models=None, cli_args=None):
    log("=== MONITORING START (Option 2: SAFE PhySO eval) ===")
    log(f"System: {sysinfo()}")
    os.makedirs(OUT_DIR, exist_ok=True)

    # Load data
    log(f"Loading dataset: {DATA_PATH}")

    # ===== Load dataset (pandas-free) =====
    from openpyxl import load_workbook

    wb = load_workbook(DATA_PATH, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    if header is None:
        raise ValueError(f"No header found in {DATA_PATH}")

    col_index = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    if TARGET_COL not in col_index:
        raise ValueError(f"Target column '{TARGET_COL}' not found in {DATA_PATH}")

    target_i = col_index[TARGET_COL]

    drop_cols = set([target_i])
    for c in ID_LIKE_COLS:
        if c in col_index:
            drop_cols.add(col_index[c])
            log(f"Dropped ID-like column: {c}")

    keep_is = [i for i in range(len(header)) if i not in drop_cols]

    X_list = []
    y_list = []
    bad_rows = 0

    for r in rows:
        if r is None:
            continue
        if all(v is None for v in r):
            continue
        yv = r[target_i] if target_i < len(r) else None
        if yv is None:
            bad_rows += 1
            continue
        try:
            y_list.append(float(yv))
        except Exception:
            bad_rows += 1
            continue

        rowx = []
        for i in keep_is:
            v = r[i] if i < len(r) else None
            if v is None:
                rowx.append(np.nan)
            else:
                try:
                    rowx.append(float(v))
                except Exception:
                    rowx.append(np.nan)
        X_list.append(rowx)

    X = np.asarray(X_list, dtype=float)
    y = np.asarray(y_list, dtype=float)

    log(f"Data loaded | X shape: {X.shape} | y shape: {y.shape} | bad_rows_skipped={bad_rows}")
    global N_SAMPLES, REGIME
    N_SAMPLES = int(X.shape[0])
    # Finalize regime now that n_samples is known
    REGIME = _select_regime(N_SAMPLES, REGIME)
    log(f"Selected regime: {REGIME} ({_regime_summary(REGIME)})")
    log(f"Detected columns | numeric={X.shape[1]} categorical=0")
    preprocessor = build_preprocessor(X)


    # Build model list
    log("Building model specs...")
    model_specs = make_model_space_list(preprocessor)
    model_specs += get_symbolic_specs(preprocessor)
    log(f"Total models configured: {len(model_specs)}")

    # ---- Filter models if requested via CLI ----
    if selected_models is not None:
        before = len(model_specs)
        model_specs = [spec for spec in model_specs if spec[0] in selected_models]
        log(f"Model filter applied | selected={sorted(selected_models)} | kept={len(model_specs)}/{before}")
        if not model_specs:
            log("[WARN] No models selected after filtering. Exiting gracefully.")
            return


    # ===== MERGE-ONLY MODE =====
    if cli_args is not None and getattr(cli_args, "merge_only", False):
        merge_dir = getattr(cli_args, "merge_dir", None) or OUT_DIR
        log(f"[MERGE] merge-only enabled | merge_dir={merge_dir}")
        summary = []
        for name, _, _, _ in model_specs:
            log(f"[MERGE] Merging fold artifacts for {name}...")
            oof_pred = merge_fold_artifacts(name, n_samples=len(y), out_dir=merge_dir, metric=TARGET_METRIC)
            if not (cli_args is not None and getattr(cli_args, "outer_fold", 0)):

                save_parity(name, y, oof_pred, OUT_DIR)
            m = metrics_dict(name, y, oof_pred)
            log(f"[MERGE] {name} - RMSE={m['RMSE']:.4f}, MAE={m['MAE']:.4f}, R2={m['R2']:.4f}")
            summary.append(m)

        # Save summary CSV
        out_summary = os.path.join(OUT_DIR, TARGET_METRIC + "_model_summary_metrics_merged.csv")
        if summary:
            summary_sorted = sorted(summary, key=lambda d: d.get("R2", float("-inf")), reverse=True)
            fieldnames = list(summary_sorted[0].keys())
            with open(out_summary, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for row in summary_sorted:
                    writer.writerow(row)
            log(f"[MERGE] Saved summary CSV: {out_summary}")
        log("[MERGE] Done.")
        return

    # ===== Outer OOF for each base model =====
    summary = []
    oof_dict = {}
    best_params_dict = {}

    for name, model, space, n_iter in model_specs:
        log(f"--- START {name} ---")
        oof_pred, best_params_list = nested_cv_oof(X, y, model, space, name, n_splits=N_SPLITS, n_iter=(n_iter or 1), outer_fold=getattr(cli_args, "outer_fold", 0), save_fold_npz=True)
        if not (cli_args is not None and getattr(cli_args, "outer_fold", 0)):

            save_parity(name, y, oof_pred, OUT_DIR)
        if not (cli_args is not None and getattr(cli_args, "outer_fold", 0)):
            m = metrics_dict(name, y, oof_pred)
            log(f"{name} - RMSE={m['RMSE']:.4f}, MAE={m['MAE']:.4f}, R2={m['R2']:.4f}")
            summary.append(m)
            oof_dict[name] = oof_pred
        best_params_dict[name] = best_params_list
        log(f"--- END {name} ---")

    # If running a single outer fold, stop here (fold artifacts already saved)
    if cli_args is not None and getattr(cli_args, "outer_fold", 0):
        log(f"Single outer fold run complete: fold={getattr(cli_args, 'outer_fold')} | skipping stacking/refit/summary")
        return

    # ===== Stacking =====
    do_stacking = True
    if cli_args is not None and getattr(cli_args, 'no_stacking', False):
        do_stacking = False
    if len(oof_dict) < 2:
        do_stacking = False

    if do_stacking:
        log("Building stacking ensemble (meta-learner on OOF predictions)...")
    if do_stacking:
            base_names = list(oof_dict.keys())
            oof_matrix = np.vstack([oof_dict[n] for n in base_names]).T
            meta = ElasticNet(alpha=0.01, l1_ratio=0.1, max_iter=20000, random_state=RANDOM_STATE)
            meta_oof = cross_val_predict(meta, oof_matrix, y, cv=RoundRobinKFold(n_splits=N_SPLITS), n_jobs=N_JOBS, verbose=1)
            save_parity("StackingEnsemble", y, meta_oof, OUT_DIR)
            m_stack = metrics_dict("StackingEnsemble", y, meta_oof)
            log(f"STACKING - RMSE={m_stack['RMSE']:.4f}, MAE={m_stack['MAE']:.4f}, R2={m_stack['R2']:.4f}")
            summary.append(m_stack)
    
    # ===== Full-data refit =====
    chosen_params = {}
    for name, model, space, _ in model_specs:
        if space is not None:
            log(f"[{name}] Fitting full-data model with aggregated best params...")
            model_full, agg_params = fit_full_with_best_params(model, best_params_dict[name], name=name)
            t0 = time.time()
            # --- final training on full data ---
            best_params = dict(agg_params)  # use aggregated params from CV

            # --- Fix types for sklearn params (RF etc.) ---
            INT_SUBSTR = ["n_estimators","max_depth","min_samples_split","min_samples_leaf","max_leaf_nodes","min_child_samples","num_leaves"]
            for k in list(best_params.keys()):
                v = best_params[k]
                if v is None:
                    continue
                if any(s in k for s in INT_SUBSTR):
                    if isinstance(v, float) and v >= 1.0:
                        best_params[k] = int(round(v))

            # Apply params and fit
            model_full.set_params(**best_params)
            model_full.fit(X, y)
            log(f"[{name}] Full-data fit completed in {time.time()-t0:.1f}s")
            chosen_params[name] = agg_params
        else:
            chosen_params[name] = {"internal_search": True}

    # ===== Save artifacts =====
    with open(os.path.join(OUT_DIR, f"{TARGET_METRIC}_chosen_params.json"), "w", encoding="utf-8") as f:
        json.dump(chosen_params, f, indent=2, ensure_ascii=False)
        log(f"Saved params JSON: {f.name}")

    # Save stacking meta only if stacking was performed
    if do_stacking:
        base_names = list(oof_dict.keys())
        meta.fit(np.vstack([oof_dict[n] for n in base_names]).T, y)
        coefs = {base_names[i]: float(meta.coef_[i]) for i in range(len(base_names))}
        with open(os.path.join(OUT_DIR, f"{TARGET_METRIC}_stacking_meta.json"), "w", encoding="utf-8") as f:
            json.dump({"intercept": float(meta.intercept_), "coefficients": coefs}, f, indent=2, ensure_ascii=False)
            log(f"Saved stacking meta JSON: {f.name}")
    else:
        log("Stacking disabled; skipping stacking meta JSON.")
    # --- Save summary metrics without pandas ---
    out_summary = os.path.join(OUT_DIR, TARGET_METRIC + "_model_summary_metrics_tuned.csv")
    if summary:
        # Sort by R2 descending if present
        try:
            summary_sorted = sorted(summary, key=lambda d: d.get("R2", float("-inf")), reverse=True)
        except Exception:
            summary_sorted = summary

        fieldnames = list(summary_sorted[0].keys())
        with open(out_summary, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in summary_sorted:
                writer.writerow(row)
        log(f"Saved summary CSV: {out_summary}")

        # Pretty print
        log("=== SUMMARY (top) ===")
        for i, row in enumerate(summary_sorted[:10], 1):
            log(f"{i:02d}. " + ", ".join([f"{k}={row.get(k)}" for k in fieldnames]))
    else:
        log("No summary rows to save.")


    log("=== MONITORING END ===")

if __name__ == "__main__":

    _AVAILABLE = ["ElasticNet","Ridge","RandomForest","MLP","XGBoost","LightGBM","PhySO","EQL", "PySR"]
    selected_models, cli_args = parse_args(_AVAILABLE)
    main(selected_models=selected_models, cli_args=cli_args)
